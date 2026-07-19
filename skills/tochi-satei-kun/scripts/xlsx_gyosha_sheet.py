# Copyright 2026 Koichi Matsuda / SignalYield Advisory
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""分割：業者用シート描画（v1.2.7、Cowork 読み込み対策）"""
from xlsx_common import *


# ===== 業者用シート =====
def _write_gyosha_sheet(wb: Workbook, ctx: dict):
    ws = wb.create_sheet("業者用")
    # 列幅を冒頭で設定（Cowork 配布層 truncate 対策。末尾の再設定もそのまま残す）
    _adjust_col_widths(ws, [14, 10, 12, 16, 12, 14, 12, 16, 10, 10, 10, 10, 12, 10])
    # 印刷範囲を冒頭で暫定設定（Cowork 配布層 truncate 対策。
    # 末尾で `r` の正確値に上書きするが、truncate された場合に備えてマージン付き暫定値を先に置く）
    ws.print_area = "A1:N200"
    # グラフ専用シートを 業者用 の直後（インデックス 1）に作成
    graph_ws = wb.create_sheet("グラフ", 1)
    # グラフシートのタイトル
    _set(graph_ws, 1, 1, "■ 附属資料",
         font=Font(name="游ゴシック", size=14, bold=True, color="FFFFFF"),
         fill=TITLE_FILL,
         align=Alignment(horizontal="left", vertical="center"))
    graph_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    graph_ws.row_dimensions[1].height = 28
    # 列幅をグラフ表示用に調整
    for col_letter in 'ABCDEFGHIJKLMN':
        graph_ws.column_dimensions[col_letter].width = 10
    # グラフ配置用の running row tracker
    ctx['_graph_ws'] = graph_ws
    ctx['_graph_row'] = 3  # タイトル(1)+空行(2)
    target = ctx["target"]
    asof = ctx["asof"]
    scope_log = ctx["scope_log"]
    rate_info = ctx["rate_info"]
    hed = ctx["hedonic"]
    cases = ctx["cases"]
    breakdown = ctx["breakdown"]
    assess = ctx["assess"]
    refs = ctx["refs"]
    standard_check = ctx["standard_check"]
    hijun_rows = ctx.get("hijun_rows", [])
    hijun_detail_rows = ctx.get("hijun_detail_rows", [])

    r = 1
    # 認証マーカー（A1）— ハルシネーション出力との判別用、INSTALL.md 検証チェックリスト参照
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(ws, r, 1, "tochi-satei-kun v1.4.3 認証出力",
         font=Font(name="游ゴシック", size=8, italic=True, color="808080"),
         align=Alignment(horizontal="left", vertical="center"))
    r += 1

    # タイトル（A2）
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(ws, r, 1, f"土地価格査定 業者用シート — {target.get('物件略号', '')} ({target['市区町村名']} {target.get('地区名', '')})",
         font=TITLE_FONT, fill=TITLE_FILL,
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[r].height = 28
    r += 2

    # ヘッダ：物件概要
    _section_header(ws, r, "■ 物件概要・スコープ")
    r += 1
    info = [
        ("査定時点", asof.isoformat()),
        ("所在", f"{target['都道府県名']} {target['市区町村名']} {target.get('地区名', '')}{target.get('丁目', '')}"),
        ("面積", f"{target['面積(㎡)']} ㎡"),
        ("最寄駅", f"{target.get('最寄駅:名称', '')} 徒歩{target.get('最寄駅:距離(分)', '')}分"),
        ("形状", target.get("土地の形状", "")),
        ("接道", f"{target.get('前面道路:種類', '')} 幅員{target.get('前面道路:幅員(m)', '')}m {target.get('前面道路:方位', '')}向"),
        ("用途地域", target.get("都市計画", "")),
        ("建ぺい率/容積率", f"{target.get('建ぺい率(%)', '')}% / {target.get('容積率(%)', '')}%"),
        ("使用事例件数", f"{scope_log['final_count']} 件 (IQR除外: {scope_log['iqr_removed']}件、市区町村単位・隣接拡張なし)"),
    ]
    for label, value in info:
        _set(ws, r, 1, label, font=LABEL_FONT, border=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        _set(ws, r, 2, value, font=VALUE_FONT, border=True)
        r += 1
    r += 1

    # 査定価格
    _section_header(ws, r, "■ 査定価格")
    r += 1
    target_area = target["面積(㎡)"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(ws, r, 1,
         _format_price_full(assess["central_total_price"], target_area),
         font=BIG_VALUE_FONT,
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[r].height = 36
    r += 1

    # 信頼度ラベル（高/中/中-低/低）：n × 自由度調整済 R² × 期待符号整合性ベース
    # 中-低 は「構造問題」（**統計的に有意な**符号反転2件以上 or adj_R² が極端に低い）に限定。
    # 非有意（p≥0.10）な符号反転はノイズ範囲内とみなしカウントしない。
    if hed["ok"]:
        n = hed["n"]
        adj_r2 = hed["adj_r2"]
        EXPECTED_NEG = ("ln_area", "walk_min", "D_shidou", "D_fukuro", "D_fuseikei")
        SIG_P_THRESHOLD = 0.10  # この p 値未満の反転のみ「有意な反転」としてカウント
        coef = hed["coefficients"]
        sign_inconsistent = sum(
            1 for name in EXPECTED_NEG
            if name in coef and coef[name]["beta"] > 0 and coef[name]["p"] < SIG_P_THRESHOLD
        )
        sign_checked = sum(1 for name in EXPECTED_NEG if name in coef)
        if sign_inconsistent >= 2 or adj_r2 < 0.3:
            reasons = []
            if sign_inconsistent >= 2:
                reasons.append(f"有意な符号反転 {sign_inconsistent}/{sign_checked} 件（p<{SIG_P_THRESHOLD}）")
            if adj_r2 < 0.3:
                reasons.append(f"adj R² = {adj_r2:.2f}（低水準）")
            conf_label = (f"モデル適合度：要注意（n = {n}, "
                          + ", ".join(reasons)
                          + " — 構造問題の可能性、要再確認）")
            conf_fill = P_HIGH_FILL
        elif n >= 20 and adj_r2 >= 0.45 and sign_inconsistent == 0:
            conf_label = (f"モデル適合度：良好（n = {n}, 自由度調整済 R² = {adj_r2:.2f}, "
                          f"有意な期待符号と全整合）")
            conf_fill = P_LOW_FILL
        else:
            reasons = []
            if n < 20:
                reasons.append(f"事例件数 n = {n} と少なめ")
            if adj_r2 < 0.45:
                reasons.append(f"adj R² = {adj_r2:.2f}（中程度）")
            if sign_inconsistent == 1:
                reasons.append(f"有意な符号反転 1/{sign_checked} 件")
            if not reasons:
                reasons.append(f"n = {n}, adj R² = {adj_r2:.2f}")
            conf_label = "モデル適合度：中程度（" + " / ".join(reasons) + "）"
            conf_fill = P_MID_FILL
    else:
        conf_label = "モデル適合度：参考情報（件数不足のため係数推定不能。顧客用シートは『参考情報』として出力）"
        conf_fill = P_HIGH_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    _set(ws, r, 1, conf_label,
         font=Font(name="游ゴシック", size=11, bold=True),
         fill=conf_fill, border=True,
         align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[r].height = 24
    r += 2

    # 2価格サマリ：採用査定価格 vs ヘドニック母集団予測（信頼度ラベル直下に配置）
    target_area_local = target["面積(㎡)"]
    central_unit = assess.get("central_unit_price")
    hed_pred = refs.get("hedonic_pred")

    _section_header(ws, r, "■ 2価格サマリ（採用査定価格 vs ヘドニック母集団予測の乖離）")
    r += 1
    for j, h in enumerate(["区分", "㎡単価", "総額", "採用査定との乖離率"]):
        _set(ws, r, j+1, h, font=LABEL_FONT,
             fill=PatternFill("solid", fgColor="D9E1F2"), border=True)
    r += 1

    # ① 土地比準（採用）
    _set(ws, r, 1, "① 土地比準（採用）",
         font=Font(name="游ゴシック", size=10, bold=True),
         fill=PRIMARY_FILL, border=True)
    if central_unit:
        _set(ws, r, 2, f"{int(central_unit):,}円",
             font=Font(name="游ゴシック", size=10, bold=True), fill=PRIMARY_FILL, border=True)
        _set(ws, r, 3, _format_jpy(central_unit * target_area_local),
             font=Font(name="游ゴシック", size=10, bold=True), fill=PRIMARY_FILL, border=True)
    _set(ws, r, 4, "—", font=VALUE_FONT, fill=PRIMARY_FILL, border=True)
    r += 1

    # ② ヘドニック母集団予測
    if hed_pred and central_unit:
        dev = (hed_pred - central_unit) / central_unit * 100
        abs_dev = abs(dev)
        if abs_dev <= 15:
            dev_fill = P_LOW_FILL
            dev_guide = "（15%以内：採用査定とヘドニック予測が概ね整合）"
        elif abs_dev <= 30:
            dev_fill = P_MID_FILL
            dev_guide = "（15〜30%：地域特性または個別事例の特殊性を確認すると良い）"
        else:
            dev_fill = P_HIGH_FILL
            dev_guide = ("※ 30%超：規範性の高い事例が母集団から外れている可能性。"
                         "事例選定と特徴量を再確認してください。")
        _set(ws, r, 1, "② ヘドニック母集団予測", font=VALUE_FONT, border=True)
        _set(ws, r, 2, f"{int(hed_pred):,}円", font=VALUE_FONT, border=True)
        _set(ws, r, 3, _format_jpy(hed_pred * target_area_local), font=VALUE_FONT, border=True)
        _set(ws, r, 4, f"{dev:+.1f}%", font=VALUE_FONT, fill=dev_fill, border=True)
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        _set(ws, r, 1, dev_guide,
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"))
        r += 1
    elif central_unit:
        _set(ws, r, 1, "② ヘドニック母集団予測", font=VALUE_FONT, border=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        _set(ws, r, 2, "（件数不足のため算出不能）", font=VALUE_FONT, fill=MISSING_FILL, border=True)
        r += 1
    r += 1

    # 価格レンジ（比準表の試算値の最大/中央/最小と一致）
    _section_header(ws, r, "■ 価格レンジ（比準表の試算値 最大／中央／最小）")
    r += 1
    rng = assess["range"]
    headers = ["区分", "総額", "㎡単価", "坪単価"]
    for j, h in enumerate(headers):
        _set(ws, r, j+1, h, font=LABEL_FONT, fill=PatternFill("solid", fgColor="D9E1F2"), border=True)
    r += 1
    for label, total, unit in [
        ("上限", rng["high_total"], rng["high_unit"]),
        ("中央", rng["central_total"], rng["central_unit"]),
        ("下限", rng["low_total"], rng["low_unit"]),
    ]:
        total_r = _round_3sig(total) if total else None
        unit_sqm_r = _round_3sig(unit) if unit else None
        unit_tsubo_r = _round_3sig(unit / 0.3025) if unit else None
        _set(ws, r, 1, label, font=LABEL_FONT, border=True)
        _set(ws, r, 2, f"{total_r:,}円" if total_r else "", font=VALUE_FONT, border=True)
        _set(ws, r, 3, f"{unit_sqm_r:,}円" if unit_sqm_r else "", font=VALUE_FONT, border=True)
        _set(ws, r, 4, f"{unit_tsubo_r:,}円" if unit_tsubo_r else "", font=VALUE_FONT, border=True)
        r += 1
    r += 1

    # 比準表（建付減価列を削除した8列構成）
    if hijun_rows:
        _insert_page_break(ws, r)
        _section_header(ws, r, "■ 比準表（標準画地の比準価格）")
        r += 1
        # 列構成（8列、建付減価削除済み）：
        # 1=事例番号, 2=取引価格, 3=事情補正, 4=時点修正,
        # 5=標準化補正, 6=地域格差, 7=試算値, 8=標準画地の価格
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        for j, h in enumerate(["事例番号", "取引価格(円/㎡)", "事情補正", "時点修正",
                               "標準化補正", "地域格差",
                               "試算値(円/㎡)", "標準画地の価格(円/㎡)"]):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=header_fill, border=True,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.row_dimensions[r].height = 32
        r += 1
        # 試算値の中央値を標準画地の価格に
        # v1.2.5: 試算値は correction.py 側で上位3桁に丸め済み。median 計算後も
        # （偶数件で平均が走るケースに備えて）上位3桁に再度丸めて精度を揃える。
        n_rows = len(hijun_rows)
        shisan_list = sorted(h["試算値"] for h in hijun_rows)
        if n_rows % 2 == 1:
            hijun_central = shisan_list[n_rows // 2]
        else:
            hijun_central = (shisan_list[n_rows // 2 - 1] + shisan_list[n_rows // 2]) / 2
        hijun_central = _round_3sig(hijun_central)

        block_start_row = r
        center_align = Alignment(horizontal="center", vertical="center")
        # 表示順を [top2, top1, top3] に並び替え（規範性の高い事例を中央に配置、視覚強調なし）
        # hijun_rows は [top1, top2, top3] の順で来る
        if len(hijun_rows) == 3:
            display_rows = [hijun_rows[1], hijun_rows[0], hijun_rows[2]]
        elif len(hijun_rows) == 2:
            display_rows = [hijun_rows[1], hijun_rows[0]]
        else:
            display_rows = hijun_rows
        for idx, h in enumerate(display_rows):
            # 色強調なし（位置で識別：中央＝規範性の高い事例）
            fill = None
            font_top = VALUE_FONT
            label_font = LABEL_FONT
            top_row = r
            bot_row = r + 1
            # 補正項目の分子/分母（鑑定書様式）
            # 時点修正：分子側（査定時点 / 事例時点）
            # 標準化補正・地域格差：分母側（100 / 事例評点 = 事例側を分母に置く慣習）
            jijo_top, jijo_bot = _hijun_top_bottom(h["事情補正"], h.get("事情補正_適用", False))
            time_top, time_bot = _hijun_top_bottom(h["時点修正"], mode="top")
            hyo_top, hyo_bot = _hijun_top_bottom(h["標準化補正"], mode="bottom")
            chi_top, chi_bot = _hijun_top_bottom(h["地域格差"], mode="bottom")
            # 上行（分子）：列 3=事情補正, 4=時点修正, 5=標準化補正, 6=地域格差
            _set(ws, top_row, 3, jijo_top, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, top_row, 4, time_top, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, top_row, 5, hyo_top, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, top_row, 6, chi_top, font=font_top, fill=fill, border=True, align=center_align)
            # 下行（分母）
            _set(ws, bot_row, 3, jijo_bot, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, bot_row, 4, time_bot, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, bot_row, 5, hyo_bot, font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, bot_row, 6, chi_bot, font=font_top, fill=fill, border=True, align=center_align)
            # 2行マージ：事例番号(1), 取引価格(2), 試算値(7)
            for col in [1, 2, 7]:
                ws.merge_cells(start_row=top_row, start_column=col,
                               end_row=bot_row, end_column=col)
            # 事例番号 = MLITデータ番号（透明性のため、人為的ラベルではない）
            case_no_str = str(h.get("事例番号", "?"))
            _set(ws, top_row, 1, case_no_str, font=label_font, fill=fill, border=True,
                 align=center_align)
            _set(ws, top_row, 2, f"{int(h['取引価格']):,}",
                 font=font_top, fill=fill, border=True, align=center_align)
            _set(ws, top_row, 7, f"{int(round(h['試算値'])):,}",
                 font=font_top, fill=fill, border=True, align=center_align)
            r += 2
        block_end_row = r - 1
        # 標準画地の価格列（8列目）を全事例マージ
        ws.merge_cells(start_row=block_start_row, start_column=8,
                       end_row=block_end_row, end_column=8)
        _set(ws, block_start_row, 8, f"{int(round(hijun_central)):,}",
             font=Font(name="游ゴシック", size=12, bold=True, color="C00000"),
             border=True,
             align=Alignment(horizontal="center", vertical="center"))
        # 注釈
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        _set(ws, r, 1,
             "※ 事例番号 = MLITデータ原本の行番号。標準画地の価格 = 3事例の試算値の中央値。"
             "各補正は「分子/分母」形式（上段=分子、下段=分母）。「100/-」は補正非該当。"
             "標準化補正＝画地条件（規模, 形状, 方位, 袋地, 不整形）、"
             "地域格差＝地域・街路・交通要因（道路幅員, 駅徒歩, 容積率, 私道, 地区平均, 駅平均）のヘドニック係数積。"
             "**中央行（2段目）＝規範性の高い事例**（top1）、上下行は検証用の類似事例。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 55
        r += 2

    # 比準表の内訳：取引事例の補修正率と地域格差率（鑑定実務標準フォーマット、9列）
    if hijun_detail_rows:
        _section_header(ws, r, "■ 比準表の内訳（取引事例の補修正率と地域格差率）",
                        end_col=9)
        r += 1

        # 2段ヘッダ：上段は「地域格差」のグループラベル
        hdr_fill = PatternFill("solid", fgColor="D9E1F2")
        col_labels_top = ["事例番号", "事情補正", "時点修正",
                          "標準化補正", "地域格差", "", "", "", ""]
        for j, h in enumerate(col_labels_top):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=hdr_fill, border=True,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        # 地域格差は4区分＋相乗積マージ（列5-9）— 相乗積=地域格差の積であることを明示
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
        # 単独列は2段マージ（縦）— 相乗積は地域格差サブヘッダ配下なので除外
        for col in [1, 2, 3, 4]:
            ws.merge_cells(start_row=r, start_column=col, end_row=r+1, end_column=col)
        r += 1
        # 下段：地域格差の4細目＋相乗積
        col_labels_bot = ["", "", "", "", "街路条件\n（総和）",
                          "交通接近条件\n（総和）", "環境条件\n（総和）", "行政的条件\n（総和）",
                          "相乗積\n（地域格差積）"]
        for j, h in enumerate(col_labels_bot):
            if h:  # 既にマージされていないセルのみ
                _set(ws, r, j+1, h, font=LABEL_FONT, fill=hdr_fill, border=True,
                     align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.row_dimensions[r-1].height = 18
        ws.row_dimensions[r].height = 30
        r += 1

        def _fmt_pct(v):
            """+X.X / ±0 / ▲X.X 形式"""
            if v is None:
                return "―"
            v = round(v, 1)
            if abs(v) < 0.05:
                return "±0"
            if v > 0:
                return f"+{v}"
            return f"▲{abs(v)}"

        def _filter_nonzero(items):
            """サブ項目のうち ±0（絶対値<0.05）を除外。
            v1.2.1: 「地区」エントリは β=0.81 と高インパクトの最重要要因なので、
            ±0でも常時表示する（事例と本物件の地区が同じことを白箱性として明示）。
            """
            keep = []
            for lbl, pct in items:
                if lbl.startswith("地区"):
                    keep.append((lbl, pct))  # 地区は常時表示
                elif abs(round(pct, 1)) >= 0.05:
                    keep.append((lbl, pct))
            return keep

        def _join_subitems(items, hide_zero=True):
            """[(label, pct), ...] を multi-line text に。±0は非表示（hide_zero=True）。"""
            if hide_zero:
                items = _filter_nonzero(items)
            if not items:
                return "標準的 ±0"
            return "\n".join(f"{lbl} {_fmt_pct(pct)}" for lbl, pct in items)

        cell_font = Font(name="游ゴシック", size=9)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 取引事例 行（事例番号で表示）— 中央に規範性の高い事例(top1)を配置
        if len(hijun_detail_rows) == 3:
            display_detail = [hijun_detail_rows[1], hijun_detail_rows[0], hijun_detail_rows[2]]
        elif len(hijun_detail_rows) == 2:
            display_detail = [hijun_detail_rows[1], hijun_detail_rows[0]]
        else:
            display_detail = hijun_detail_rows
        for idx, d in enumerate(display_detail):
            fill = None
            f = cell_font
            _set(ws, r, 1, d.get("事例番号", "?"), font=f, fill=fill, border=True, align=center_align)
            jijo_lbl, jijo_pct = d.get("事情補正", ("正常", 0.0))
            _set(ws, r, 2, f"{jijo_lbl}\n{_fmt_pct(jijo_pct)}", font=f, fill=fill, border=True, align=center_align)
            _set(ws, r, 3, _fmt_pct(d.get("時点修正_pct", 0)), font=f, fill=fill, border=True, align=center_align)
            std_items = _filter_nonzero(d.get("規模", []) + d.get("画地", []))
            std_text = "\n".join(f"{lbl} {_fmt_pct(pct)}" for lbl, pct in std_items) if std_items else "標準的 ±0"
            std_text += f"\n総和 {_fmt_pct(d.get('標準化補正_総和', 0))}"
            _set(ws, r, 4, std_text, font=f, fill=fill, border=True, align=center_align)
            street_text = _join_subitems(d.get("街路", []))
            street_text += f"\n総和 {_fmt_pct(d.get('街路_総和', 0))}"
            _set(ws, r, 5, street_text, font=f, fill=fill, border=True, align=center_align)
            tr_text = _join_subitems(d.get("交通接近", []))
            tr_text += f"\n総和 {_fmt_pct(d.get('交通接近_総和', 0))}"
            _set(ws, r, 6, tr_text, font=f, fill=fill, border=True, align=center_align)
            env_text = _join_subitems(d.get("環境", []))
            env_text += f"\n総和 {_fmt_pct(d.get('環境_総和', 0))}"
            _set(ws, r, 7, env_text, font=f, fill=fill, border=True, align=center_align)
            adm_text = _join_subitems(d.get("行政", []))
            adm_text += f"\n総和 {_fmt_pct(d.get('行政_総和', 0))}"
            _set(ws, r, 8, adm_text, font=f, fill=fill, border=True, align=center_align)
            _set(ws, r, 9, d.get("相乗積", 100), font=f, fill=fill, border=True, align=center_align)
            ws.row_dimensions[r].height = 70
            r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        koji_rate = rate_info.get("rate", 0) or 0
        _set(ws, r, 1,
             f"※ 時点修正率査定根拠：地価公示の年次変動率を参考に、地域の地価動向を分析の上、"
             f"年率 {koji_rate*100:+.1f}% で査定（{rate_info.get('method','')}, n = {rate_info.get('n_points', 0)} 地点）。"
             "各補正率は %-point 表記、相乗積は 100 を基準とする指数。"
             "**中央行（2段目）＝規範性の高い事例**（top1）、上下行は検証用の類似事例。"
             "標準化補正の細目（規模・形状・方位）と地域格差の細目（街路・交通接近・環境・行政）は、"
             "対応するヘドニック係数を反映。±0 の細目は非表示。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 50
        r += 2

    # ■ 個別格差 + 査定価格の算定（業者用：縦並び個別格差 + 横並び算定式）
    if hijun_rows:
        # 規範性の高い事例（top1）の個別格差を採用
        primary_h_gy = next((h for h in hijun_rows if h.get("順位") == "規範性の高い事例"),
                            hijun_rows[0])
        center_align_gy = Alignment(horizontal="center", vertical="center")
        right_align_gy = Alignment(horizontal="right", vertical="center")

        # 標準画地の試算値（= 標準画地の価格 = 3事例の試算値の中央値）— 既に比準表で計算済み
        hijun_central_val = int(round(hijun_central))

        # 個別格差ブロック開始行
        kobetsu_start_row = r
        _set(ws, r, 1, "■ 個別格差",
             font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
             fill=SECTION_FILL,
             align=Alignment(horizontal="left", vertical="center"))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        # 査定価格の算定ヘッダ（右側）
        _set(ws, r, 4, "■ 査定価格の算定",
             font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
             fill=SECTION_FILL,
             align=Alignment(horizontal="left", vertical="center"))
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        r += 1

        def _fmt_kobetsu_v(v):
            iv = round(v, 1)
            return int(iv) if iv == int(iv) else iv

        # 個別格差 縦表（列A=ラベル、列B=値）
        # v1.2.1: target の属性をラベル括弧に転記（Style B）、target が中間画地の場合は角地行を非表示

        # 角地（target に角地補正率 > 0 が明示入力された場合のみ表示）
        target_kado_val_gy = primary_h_gy.get("個別格差_角地", 0)
        kado_row_gy = None  # 非表示の場合は None
        if abs(round(target_kado_val_gy, 1)) >= 0.05:
            _set(ws, r, 1, "角地（角地）", font=LABEL_FONT, border=True, align=center_align_gy)
            _set(ws, r, 2, _fmt_kobetsu_v(target_kado_val_gy),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            kado_row_gy = r
            r += 1

        # 算定式の視覚アンカー行（角地が非表示なら方位の行に揃える）
        first_kobetsu_row = kado_row_gy if kado_row_gy is not None else r

        # 方位（target の方位を表示ラベルに）
        target_dir_gy = str(target.get("前面道路:方位", "")).strip()
        houi_label_gy = f"方位（{target_dir_gy}）" if target_dir_gy else "方位"
        houi_val = _fmt_kobetsu_v(primary_h_gy.get("個別格差_方位", 0))
        _set(ws, r, 1, houi_label_gy, font=LABEL_FONT, border=True, align=center_align_gy)
        _set(ws, r, 2, houi_val, font=VALUE_FONT, border=True, align=center_align_gy)
        houi_row_gy = r
        r += 1

        # 不整形（target の土地形状を表示ラベルに、v1.2.1）
        target_shape_gy = str(target.get("土地の形状", "")).strip()
        fusei_label_gy = f"不整形（{target_shape_gy}）" if target_shape_gy else "不整形"
        fusei_val = _fmt_kobetsu_v(primary_h_gy.get("個別格差_不整形", 0))
        _set(ws, r, 1, fusei_label_gy, font=LABEL_FONT, border=True, align=center_align_gy)
        _set(ws, r, 2, fusei_val, font=VALUE_FONT, border=True, align=center_align_gy)
        fusei_row_gy = r
        r += 1

        # 総和（Excel関数式）— 表示中の格差行のみを積算
        _set(ws, r, 1, "総和", font=LABEL_FONT, fill=PatternFill("solid", fgColor="FFF2CC"),
             border=True, align=center_align_gy)
        factor_refs_gy = []
        if kado_row_gy is not None:
            factor_refs_gy.append(f"(100+B{kado_row_gy})/100")
        factor_refs_gy.append(f"(100+B{houi_row_gy})/100")
        factor_refs_gy.append(f"(100+B{fusei_row_gy})/100")
        soan_formula_gy = "=" + "*".join(factor_refs_gy) + "*100"
        soan_cell_gy = ws.cell(row=r, column=2, value=soan_formula_gy)
        soan_cell_gy.font = Font(name="游ゴシック", size=10, bold=True)
        soan_cell_gy.border = BORDER
        soan_cell_gy.alignment = center_align_gy
        soan_cell_gy.number_format = "0.00"
        soan_cell_gy.fill = PatternFill("solid", fgColor="FFF2CC")
        soan_row_gy = r
        r += 1

        # ====== 査定価格の算定 行（個別格差ブロックの隣、first_kobetsu_row 行から横並び）======
        # 方位・不整形・明示角地は正本価格に適用済み。ここは検算用に100%を掛ける。
        # v1.2.1: 視覚アンカーは first_kobetsu_row（角地非表示時は方位の行）

        # 試算値 (D列、first_kobetsu_row 行)
        _set(ws, first_kobetsu_row, 4, hijun_central_val,
             font=Font(name="游ゴシック", size=12, bold=True),
             border=True, align=center_align_gy, number_format="#,##0")
        _set(ws, first_kobetsu_row+1, 4, "標準画地の試算値(円/㎡)",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=center_align_gy)

        # × 演算子 (E列)
        _set(ws, first_kobetsu_row, 5, "×",
             font=Font(name="游ゴシック", size=14, bold=True),
             align=center_align_gy)

        # 総和/100 (F列) — 分子=B{soan_row_gy}, 分母=100 を 2行縦に表示
        soan_ref_cell = ws.cell(row=first_kobetsu_row, column=6, value=f"=B{soan_row_gy}")
        soan_ref_cell.font = Font(name="游ゴシック", size=11, bold=True)
        soan_ref_cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=Side(border_style="thin", color="000000"))
        soan_ref_cell.alignment = center_align_gy
        soan_ref_cell.number_format = "0.00"
        denom_cell = ws.cell(row=first_kobetsu_row+1, column=6, value=100)
        denom_cell.font = Font(name="游ゴシック", size=11, bold=True)
        denom_cell.border = Border(left=THIN, right=THIN, top=Side(border_style="thin", color="000000"), bottom=THIN)
        denom_cell.alignment = center_align_gy

        # ≒ (G列)
        _set(ws, first_kobetsu_row, 7, "≒",
             font=Font(name="游ゴシック", size=14, bold=True),
             align=center_align_gy)

        # 案件査定価格 (H列) — Excel関数式
        anken_inner = f"D{first_kobetsu_row}*B{soan_row_gy}"
        anken_formula_gy = f"=ROUND({anken_inner},-(LEN(INT({anken_inner}))-3))/100"
        anken_cell_gy = ws.cell(row=first_kobetsu_row, column=8, value=anken_formula_gy)
        anken_cell_gy.font = Font(name="游ゴシック", size=14, bold=True, color="C00000")
        anken_cell_gy.border = BORDER
        anken_cell_gy.alignment = center_align_gy
        anken_cell_gy.number_format = "#,##0"
        _set(ws, first_kobetsu_row+1, 8, "案件査定価格(円/㎡)",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=center_align_gy)

        # 個別格差 + 査定価格の算定 ブロックの後は r が総和の次に進んでいる
        r += 1  # 空行

        # 注釈
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        _set(ws, r, 1,
             "※ 個別格差は規範性の高い事例（top1）と本物件の差から算出。"
             "**角地補正は業者の入力値（デフォルト 0%）**。"
             "MLITデータに角地情報が無いためヘドニックで推定不能 → 白箱ポリシー上、自動値は与えず業者判断に委ねる。"
             "方位・不整形補正はヘドニック係数 β（dir_score, D_fuseikei）に基づく "
             "exp(β×(本物件 − 事例)) − 1 として正本価格へ1回だけ反映済み。"
             "個別格差の総和欄は二重計上を避けるため100%とし、"
             "案件査定価格 = 正本補正後単価（表示上は試算値）を上位3桁四捨五入。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 45
        r += 2

        # ■ 取引事例の概要（横並び、3事例の詳細データ）
        _insert_page_break(ws, r)
        _section_header(ws, r, "■ 取引事例の概要", end_col=12)
        r += 1
        gaiyo_headers = ["事例番号", "取引㎡単価", "取引時点", "地区", "最寄り駅",
                         "駅距離(分)", "道路", "道路幅員(m)", "方位", "形状",
                         "地積(㎡)", "用途地域", "容積率(%)"]
        hdr_fill_g = PatternFill("solid", fgColor="D9E1F2")
        # 13列に拡張、セクションヘッダのマージも更新
        ws.unmerge_cells(start_row=r-1, start_column=1, end_row=r-1, end_column=12)
        ws.merge_cells(start_row=r-1, start_column=1, end_row=r-1, end_column=13)
        for j, h in enumerate(gaiyo_headers):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=hdr_fill_g, border=True,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.row_dimensions[r].height = 30
        r += 1

        # 表示順は比準表と整合：[top2, top1, top3]（中央=規範性の高い事例）
        if len(hijun_rows) == 3:
            gaiyo_display = [hijun_rows[1], hijun_rows[0], hijun_rows[2]]
        elif len(hijun_rows) == 2:
            gaiyo_display = [hijun_rows[1], hijun_rows[0]]
        else:
            gaiyo_display = hijun_rows

        def _fmt_or_dash(v, kind="num"):
            """欠損値は ― で表示。"""
            if v is None or v == "" or (isinstance(v, float) and pd.isna(v)):
                return "―"
            if kind == "int":
                try:
                    return int(v)
                except (TypeError, ValueError):
                    return str(v)
            if kind == "num":
                try:
                    return f"{int(round(float(v))):,}"
                except (TypeError, ValueError):
                    return str(v)
            return str(v)

        for h in gaiyo_display:
            _set(ws, r, 1, str(h.get("事例番号", "?")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 2, _fmt_or_dash(h.get("取引価格"), "num"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 3, str(h.get("取引四半期", "") or h.get("取引時点", "")),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 4, str(h.get("地区", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 5, str(h.get("最寄駅", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 6, _fmt_or_dash(h.get("駅距離"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 7, str(h.get("道路種別", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 8, _fmt_or_dash(h.get("道路幅員"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 9, str(h.get("方位", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 10, str(h.get("形状", "")) or "—", font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 11, _fmt_or_dash(h.get("面積"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 12, str(h.get("用途地域", "")), font=VALUE_FONT, border=True, align=center_align_gy)
            _set(ws, r, 13, _fmt_or_dash(h.get("容積率_pct"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_gy)
            r += 1

        # 注釈
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        _set(ws, r, 1,
             "※ 事例番号 = MLITデータ原本の行番号。**中央行＝規範性の高い事例**（top1）。"
             "取引時点は四半期表記（例：2025年第2四半期）。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 30
        r += 2

    # ■ 公示価格の概要（業者用：取引事例の概要と同様の横並びテーブル）
    # 地域標準価格チェックで選定された公示標準地の詳細属性を表示
    koji_points_for_summary = standard_check.get("selected_points", []) if standard_check else []
    if koji_points_for_summary:
        _section_header(ws, r, "■ 公示価格の概要", end_col=12)
        r += 1
        koji_headers = ["公示番号", "公示価格(円/㎡)", "所在", "地区", "最寄駅",
                        "駅距離(m)", "道路", "道路幅員(m)", "方位", "形状",
                        "地積(㎡)", "用途地域", "容積率(%)"]
        # 13列に拡張、セクションヘッダのマージも更新
        ws.unmerge_cells(start_row=r-1, start_column=1, end_row=r-1, end_column=12)
        ws.merge_cells(start_row=r-1, start_column=1, end_row=r-1, end_column=13)
        hdr_fill_kj = PatternFill("solid", fgColor="D9E1F2")
        for j, h in enumerate(koji_headers):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=hdr_fill_kj, border=True,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
        ws.row_dimensions[r].height = 30
        r += 1

        def _fmt_or_dash_k(v, kind="num"):
            if v is None or v == "" or v == "_":
                return "―"
            try:
                if isinstance(v, float) and pd.isna(v):
                    return "―"
            except (TypeError, ValueError):
                pass
            if kind == "int":
                try:
                    return int(float(v))
                except (TypeError, ValueError):
                    return str(v)
            if kind == "num":
                try:
                    return f"{int(round(float(v))):,}"
                except (TypeError, ValueError):
                    return str(v)
            return str(v)

        center_align_kj = Alignment(horizontal="center", vertical="center")
        # 通常は1地点（場所による価格水準差を排除するため類似度スコアで絞込み）
        for pt in koji_points_for_summary[:5]:
            _set(ws, r, 1, _short_koji_id(str(pt.get("id", ""))),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 2, _fmt_or_dash_k(pt.get("price_at_asof"), "num"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 3, _short_koji_addr(str(pt.get("address", "")),
                                             str(pt.get("district", ""))),
                 font=VALUE_FONT, border=True,
                 align=Alignment(horizontal="center", vertical="center"))
            _set(ws, r, 4, str(pt.get("district", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 5, str(pt.get("station", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 6, _fmt_or_dash_k(pt.get("station_dist_m"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 7, str(pt.get("road_type", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 8, _fmt_or_dash_k(pt.get("road_width"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 9, str(pt.get("road_dir", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 10, _koji_shape_label(pt.get("frontage_ratio"),
                                              pt.get("depth_ratio")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 11, _fmt_or_dash_k(pt.get("area_sqm"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 12, str(pt.get("zoning", "")),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            _set(ws, r, 13, _fmt_or_dash_k(pt.get("floor_area_ratio"), "int"),
                 font=VALUE_FONT, border=True, align=center_align_kj)
            r += 1

        # 注釈
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        _set(ws, r, 1,
             "※ 公示番号 = 「市区町村-連番」形式（例：世田谷-50 = 13112-000-050）。"
             "公示価格は査定時点へ線形補間済み。"
             "形状は間口比率(L01_036)・奥行比率(L01_037)から推定（最大比 ≤1.5: 整形、≤2.5: やや細長、それ以上: 細長）。",
             font=Font(name="游ゴシック", size=9, italic=True, color="595959"),
             align=Alignment(wrap_text=True, vertical="top"))
        ws.row_dimensions[r].height = 40
        r += 2

    # 公示地価の時系列推移（折れ線グラフ）— 時点修正に使用した標準地と整合
    koji_ts_obj = ctx.get("koji_timeseries", {})
    if isinstance(koji_ts_obj, dict):
        koji_ts = koji_ts_obj.get("data", [])
        koji_label = koji_ts_obj.get("label", "")
    else:
        koji_ts = koji_ts_obj
        koji_label = ""
    if len(koji_ts) >= 2:
        ts_header_row = r
        _set(ws, r, 1,
             f"▼ {koji_label} の直近5年間の価格推移",
             font=Font(name="游ゴシック", size=10, bold=True, color="595959"))
        r += 1
        for j, h in enumerate(["評価年", "平均単価 (円/㎡)"]):
            _set(ws, r, j+1, h, font=LABEL_FONT,
                 fill=PatternFill("solid", fgColor="D9E1F2"), border=True,
                 align=Alignment(horizontal="center", vertical="center"))
        r += 1
        ts_data_start = r
        for pt in koji_ts:
            _set(ws, r, 1, pt["year"], font=VALUE_FONT, border=True,
                 align=Alignment(horizontal="center", vertical="center"))
            _set(ws, r, 2, pt["price"], font=VALUE_FONT, border=True,
                 number_format='#,##0',
                 align=Alignment(horizontal="right", vertical="center"))
            r += 1
        ts_data_end = r - 1

        # 折れ線グラフ（内部タイトルなし。グラフシートの section header に統一）
        line = LineChart()
        line.title = None
        line.legend = None
        line.height = 7
        line.width = 14
        data_ref = Reference(ws, min_col=2, min_row=ts_data_start, max_col=2, max_row=ts_data_end)
        cats_ref = Reference(ws, min_col=1, min_row=ts_data_start, max_col=1, max_row=ts_data_end)
        line.add_data(data_ref, titles_from_data=False)
        line.set_categories(cats_ref)
        line.y_axis.title = "単価 (円/㎡)"
        line.x_axis.title = "評価年"
        # Y軸範囲と目盛を明示設定（折れ線が中央に来るよう、かつ目盛を表示）
        prices = [pt["price"] for pt in koji_ts]
        y_min = min(prices)
        y_max = max(prices)
        if y_max > y_min:
            margin = (y_max - y_min) * 0.3
        else:
            margin = y_max * 0.05
        axis_min = max(0, y_min - margin)
        axis_max = y_max + margin
        line.y_axis.scaling.min = axis_min
        line.y_axis.scaling.max = axis_max
        # 目盛間隔を 5分割に
        line.y_axis.majorUnit = (axis_max - axis_min) / 5
        line.y_axis.delete = False
        line.y_axis.majorTickMark = 'out'
        line.y_axis.number_format = '#,##0'
        # データラベル：年 + 値 を併記
        dl_line = DataLabelList()
        dl_line.showVal = True
        dl_line.showCatName = True
        dl_line.showSerName = False
        dl_line.showLegendKey = False
        dl_line.position = 't'
        dl_line.separator = '\n'
        line.dataLabels = dl_line
        line.x_axis.delete = False
        line.series[0].smooth = False
        # グラフはグラフ専用シートに配置
        graph_ws_ref = ctx.get('_graph_ws')
        if graph_ws_ref is not None:
            gr = ctx.get('_graph_row', 3)
            # セクション見出し
            _set(graph_ws_ref, gr, 1,
                 f"■ {koji_label} の直近5年間の価格推移",
                 font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
                 fill=SECTION_FILL,
                 align=Alignment(horizontal="left", vertical="center"))
            graph_ws_ref.merge_cells(start_row=gr, start_column=1, end_row=gr, end_column=14)
            graph_ws_ref.add_chart(line, f"A{gr+1}")
            ctx['_graph_row'] = gr + 17  # line chart 7cm ≈ 15行 + buffer
        else:
            ws.add_chart(line, f"D{ts_header_row}")
    r += 1

    # ヘドニック回帰サマリ + β符号チェック（末尾：技術詳細・係数全開示の参考情報）
    _insert_page_break(ws, r)
    _section_header(ws, r, "■ ヘドニック回帰サマリ（係数全開示・参考情報）")
    r += 1
    if hed["ok"]:
        _set(ws, r, 1, f"サンプル数 n = {hed['n']}", font=VALUE_FONT)
        _set(ws, r, 3, f"R² = {hed['r2']:.3f}", font=VALUE_FONT)
        _set(ws, r, 5, f"自由度調整済 R² = {hed['adj_r2']:.3f}", font=VALUE_FONT)
        r += 1
        for j, h in enumerate(["特徴量", "推定値 β", "標準誤差", "p値", "有意性"]):
            _set(ws, r, j+1, h, font=LABEL_FONT, fill=PatternFill("solid", fgColor="D9E1F2"), border=True)
        r += 1
        coef_data_start = r  # 表での開始行（チャートのカテゴリ範囲開始）
        for name, c in hed["coefficients"].items():
            if name == "const":
                continue  # グラフから定数項は除外
            p = c["p"]
            if p < 0.05: fill = P_LOW_FILL; sig = "** (p<0.05)"
            elif p < 0.10: fill = P_MID_FILL; sig = "*  (p<0.10)"
            else: fill = P_HIGH_FILL; sig = "ns"
            # 表示用：業者用シート column 1〜5（簡潔ラベル + 数値）
            _set(ws, r, 1, c["label"], font=VALUE_FONT, border=True)
            _set(ws, r, 2, float(c['beta']), font=VALUE_FONT, border=True,
                 number_format='+0.0000;-0.0000;0.0000')
            _set(ws, r, 3, float(c['se']), font=VALUE_FONT, border=True,
                 number_format='0.0000')
            _set(ws, r, 4, float(p), font=VALUE_FONT, border=True, fill=fill,
                 number_format='0.0000')
            _set(ws, r, 5, sig, font=VALUE_FONT, border=True, fill=fill)
            r += 1
        coef_data_end = r - 1
        # 定数項を末尾に追加（参考表示、グラフ対象外）
        if "const" in hed["coefficients"]:
            c = hed["coefficients"]["const"]
            p = c["p"]
            if p < 0.05: fill = P_LOW_FILL; sig = "** (p<0.05)"
            elif p < 0.10: fill = P_MID_FILL; sig = "*  (p<0.10)"
            else: fill = P_HIGH_FILL; sig = "ns"
            _set(ws, r, 1, c["label"], font=VALUE_FONT, border=True)
            _set(ws, r, 2, float(c['beta']), font=VALUE_FONT, border=True,
                 number_format='+0.0000;-0.0000;0.0000')
            _set(ws, r, 3, float(c['se']), font=VALUE_FONT, border=True,
                 number_format='0.0000')
            _set(ws, r, 4, float(p), font=VALUE_FONT, border=True, fill=fill,
                 number_format='0.0000')
            _set(ws, r, 5, sig, font=VALUE_FONT, border=True, fill=fill)
            r += 1
        r += 1

        # ヘドニック係数 棒グラフ（白箱AVM の象徴：全特徴量の β を可視化）
        bar = BarChart()
        bar.type = "bar"  # 横向き棒グラフ（特徴量名が長いので）
        bar.style = 11
        bar.title = None  # グラフシートの section header に統一
        bar.legend = None
        bar.height = 10  # cm
        bar.width = 16   # cm
        # チャートは業者用シートの column 1（特徴量名）と column 2（β値）を直接参照
        data_ref = Reference(ws, min_col=2, min_row=coef_data_start,
                             max_col=2, max_row=coef_data_end)
        cats_ref = Reference(ws, min_col=1, min_row=coef_data_start,
                             max_col=1, max_row=coef_data_end)
        bar.add_data(data_ref, titles_from_data=False)
        bar.set_categories(cats_ref)
        bar.y_axis.title = None
        bar.x_axis.title = "係数 β（負＝単価↓、正＝単価↑）"
        bar.y_axis.delete = False
        # データラベル：カテゴリ名 + 値 を表示（Y軸ラベルが Excel で表示されない問題への対処）
        dl = DataLabelList()
        dl.showVal = True
        dl.showCatName = True   # 棒の右側に「面積 +0.0729」形式で表示
        dl.showSerName = False
        dl.showLegendKey = False
        dl.position = 'outEnd'
        dl.separator = ' '
        bar.dataLabels = dl
        # グラフはグラフ専用シートに配置
        graph_ws_ref = ctx.get('_graph_ws')
        if graph_ws_ref is not None:
            gr = ctx.get('_graph_row', 3)
            _set(graph_ws_ref, gr, 1,
                 "■ ヘドニック回帰係数 β（単価への影響度・対数空間）",
                 font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
                 fill=SECTION_FILL,
                 align=Alignment(horizontal="left", vertical="center"))
            graph_ws_ref.merge_cells(start_row=gr, start_column=1, end_row=gr, end_column=14)
            graph_ws_ref.add_chart(bar, f"A{gr+1}")
            ctx['_graph_row'] = gr + 24  # bar chart 10cm ≈ 22行 + buffer
        else:
            anchor_cell = f"G{coef_data_start}"
            ws.add_chart(bar, anchor_cell)

        # β符号チェック（期待符号 vs 実際の符号）
        EXPECTED_SIGNS = {
            "ln_area": ("負", "面積大→単価下落"),
            "walk_min": ("負", "駅遠→単価下落"),
            "D_shidou": ("負", "私道→減価"),
            "D_fukuro": ("負", "袋地→減価"),
            "D_fuseikei": ("負", "不整形→減価"),
        }
        coef = hed["coefficients"]
        sign_check_label = Font(name="游ゴシック", size=10, bold=True, color="595959")
        _set(ws, r, 1, "▼ β符号チェック（期待符号 vs 実際）", font=sign_check_label)
        r += 1
        for j, h in enumerate(["特徴量", "期待符号", "実際 β", "整合", "経済的解釈"]):
            _set(ws, r, j+1, h, font=LABEL_FONT,
                 fill=PatternFill("solid", fgColor="D9E1F2"), border=True)
        r += 1
        any_significant_inconsistent = False
        for name, (expected, interpretation) in EXPECTED_SIGNS.items():
            if name not in coef:
                continue
            beta = coef[name]["beta"]
            p = coef[name]["p"]
            is_neg_expected = (expected == "負")
            is_consistent = (is_neg_expected and beta < 0) or (not is_neg_expected and beta > 0)
            is_significant = p < 0.10  # p<0.10 で統計的に有意
            if not is_consistent and is_significant:
                any_significant_inconsistent = True
                mark = "× 反転（有意・要確認）"
                ok_fill = P_HIGH_FILL
            elif not is_consistent:
                mark = "△ 反転（非有意・ノイズ範囲）"
                ok_fill = P_MID_FILL
            else:
                mark = "○ 整合"
                ok_fill = P_LOW_FILL
            _set(ws, r, 1, coef[name]["label"], font=VALUE_FONT, border=True)
            _set(ws, r, 2, expected, font=VALUE_FONT, border=True)
            _set(ws, r, 3, f"{beta:+.4f}", font=VALUE_FONT, border=True)
            _set(ws, r, 4, mark, font=VALUE_FONT, border=True, fill=ok_fill)
            _set(ws, r, 5, interpretation, font=VALUE_FONT, border=True)
            r += 1
        if any_significant_inconsistent:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            _set(ws, r, 1,
                 "※ 統計的に有意な符号反転（p<0.10）は外れ値・特徴量不足・地区特性などの構造問題の可能性。事例を再確認してください。"
                 "非有意な反転（△）はノイズ範囲内のため実害なし。",
                 font=VALUE_FONT, fill=WARN_FILL)
            r += 1
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        _set(ws, r, 1, f"※ {hed['skip_reason']}（類似度ベース集約に降格）",
             font=VALUE_FONT, fill=WARN_FILL)
        r += 1
    r += 1

    # 散布図：全事例の駅距離 vs 時点修正後単価（査定価格と top3 を強調）
    adjusted_full = ctx.get("adjusted_full")
    if adjusted_full is not None and len(adjusted_full) >= 10:
        scatter_header_row = r
        _set(ws, r, 1, "▼ 散布図：駅距離 vs 単価（全事例・比較事例・査定価格）",
             font=Font(name="游ゴシック", size=10, bold=True, color="595959"))
        r += 1

        # データを 列 18-23 に書き込み（右側に隠れる、列幅も狭く）
        # 18=全事例 X, 19=全事例 Y, 20=top3 X, 21=top3 Y, 22=対象 X, 23=対象 Y
        scoped_data = []
        for _, rw in adjusted_full.iterrows():
            walk = rw.get("walk_min")
            price = rw.get("adjusted_unit_price") if "adjusted_unit_price" in rw else rw.get("unit_price")
            try:
                if walk is not None and not pd.isna(walk) and price is not None and not pd.isna(price):
                    scoped_data.append((float(walk), float(price)))
            except (TypeError, ValueError):
                pass

        # ヘッダ行（列幅縮小用に色だけつけて値は書かない）
        scatter_data_start = r
        for i, (walk, price) in enumerate(scoped_data):
            ws.cell(row=scatter_data_start + i, column=18, value=walk)
            ws.cell(row=scatter_data_start + i, column=19, value=price)
        scatter_scoped_end = scatter_data_start + len(scoped_data) - 1

        # top3 を column 20-21 に
        top3_data = []
        for _, rw in cases.iterrows():
            walk = rw.get("walk_min")
            price = rw.get("corrected_unit_price")
            if price is None or pd.isna(price):
                price = rw.get("adjusted_unit_price") or rw.get("unit_price")
            try:
                if walk is not None and not pd.isna(walk) and price is not None and not pd.isna(price):
                    top3_data.append((float(walk), float(price)))
            except (TypeError, ValueError):
                pass
        for i, (walk, price) in enumerate(top3_data):
            ws.cell(row=scatter_data_start + i, column=20, value=walk)
            ws.cell(row=scatter_data_start + i, column=21, value=price)
        scatter_top3_end = scatter_data_start + max(0, len(top3_data) - 1)

        # 査定価格を column 22-23 に（1点）
        target_walk = target.get("最寄駅:距離(分)")
        target_price = assess.get("central_unit_price")
        target_present = False
        if target_walk is not None and target_price is not None:
            try:
                ws.cell(row=scatter_data_start, column=22, value=float(target_walk))
                ws.cell(row=scatter_data_start, column=23, value=float(target_price))
                target_present = True
            except (TypeError, ValueError):
                pass

        # データ列の幅を狭く（右に隠す）
        for col_letter in ['R', 'S', 'T', 'U', 'V', 'W']:
            ws.column_dimensions[col_letter].width = 3

        # 散布図（マーカーのみ、線なし）
        from openpyxl.chart.marker import Marker
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        from openpyxl.drawing.fill import ColorChoice

        sc = ScatterChart()
        sc.title = None  # グラフシートの section header に統一
        sc.style = 13
        sc.height = 9
        sc.width = 16
        sc.scatterStyle = "marker"  # 線なし、マーカーのみ
        sc.x_axis.title = "最寄駅徒歩(分)"
        sc.y_axis.title = "時点修正後単価 (円/㎡)"

        def _styled_series(y_ref, x_ref, title, color, size, symbol='circle'):
            ser = Series(y_ref, x_ref, title=title)
            # 線を非表示
            ser.graphicalProperties = GraphicalProperties()
            ser.graphicalProperties.line = LineProperties(noFill=True)
            # マーカー設定
            mk = Marker(symbol=symbol, size=size)
            mk.graphicalProperties = GraphicalProperties(solidFill=color)
            mk.graphicalProperties.line = LineProperties(solidFill=color)
            ser.marker = mk
            return ser

        # Series 1: 全事例（青小マーカー）
        if scoped_data:
            x_all = Reference(ws, min_col=18, min_row=scatter_data_start, max_col=18, max_row=scatter_scoped_end)
            y_all = Reference(ws, min_col=19, min_row=scatter_data_start, max_col=19, max_row=scatter_scoped_end)
            sc.series.append(_styled_series(y_all, x_all, "全事例", "4472C4", 4, 'circle'))
        # Series 2: top3（赤大マーカー）
        if top3_data:
            x_t3 = Reference(ws, min_col=20, min_row=scatter_data_start, max_col=20, max_row=scatter_top3_end)
            y_t3 = Reference(ws, min_col=21, min_row=scatter_data_start, max_col=21, max_row=scatter_top3_end)
            sc.series.append(_styled_series(y_t3, x_t3, "比較事例top3", "C00000", 9, 'diamond'))
        # Series 3: 査定価格（緑★大マーカー）
        if target_present:
            x_tg = Reference(ws, min_col=22, min_row=scatter_data_start, max_col=22, max_row=scatter_data_start)
            y_tg = Reference(ws, min_col=23, min_row=scatter_data_start, max_col=23, max_row=scatter_data_start)
            sc.series.append(_styled_series(y_tg, x_tg, "査定価格", "00B050", 14, 'star'))

        # グラフはグラフ専用シートに配置
        graph_ws_ref = ctx.get('_graph_ws')
        if graph_ws_ref is not None:
            gr = ctx.get('_graph_row', 3)
            _set(graph_ws_ref, gr, 1,
                 "■ 散布図：駅距離 vs 単価（全事例青、比較事例top3赤、査定価格緑）",
                 font=Font(name="游ゴシック", size=11, bold=True, color="FFFFFF"),
                 fill=SECTION_FILL,
                 align=Alignment(horizontal="left", vertical="center"))
            graph_ws_ref.merge_cells(start_row=gr, start_column=1, end_row=gr, end_column=14)
            graph_ws_ref.add_chart(sc, f"A{gr+1}")
            ctx['_graph_row'] = gr + 21  # scatter chart 9cm ≈ 19行 + buffer
        else:
            ws.add_chart(sc, f"G{scatter_header_row}")
        # スキップして次のセクションへ進む（データ書き込みは右側列なので r は変えない）
        r += 2

    _adjust_col_widths(ws, [14, 10, 12, 16, 12, 14, 12, 16, 10, 10, 10, 10, 12, 10])
    # 印刷範囲を明示指定（散布図用の隠しデータ R-W 列 / row 1671 までを印刷から除外）
    ws.print_area = f"A1:N{r}"


# ===== 顧客用シート =====
