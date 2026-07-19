# Copyright 2026 Koichi Matsuda / SignalYield Advisory
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""xlsx 生成の共通モジュール（v1.2.7 で分割）：
- スタイル定数（フォント・塗りつぶし・罫線）
- 共通ヘルパー関数（_set, _section_header, _format_jpy 等）
- 短縮ラベル辞書（SHORT_FEATURE_LABELS）

xlsx_writer.py / xlsx_gyosha_sheet.py / xlsx_kokyaku_sheet.py から
`from xlsx_common import *` で取り込む。アンダースコア接頭辞の
ヘルパーも公開対象とするため __all__ を明示的に列挙している。
"""
import math
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList

__all__ = [
    # 標準ライブラリの再エクスポート（各シートモジュールが使う）
    "math", "date", "Path", "pd",
    "Workbook", "Font", "Alignment", "PatternFill", "Border", "Side",
    "get_column_letter",
    "BarChart", "LineChart", "ScatterChart", "Reference", "Series",
    "DataLabelList",
    # 短縮ラベル
    "SHORT_FEATURE_LABELS",
    # forbidden_words
    "assert_clean",
    # スタイル定数
    "TITLE_FONT", "TITLE_FILL", "SECTION_FONT", "SECTION_FILL",
    "LABEL_FONT", "VALUE_FONT", "BIG_VALUE_FONT",
    "WARN_FILL", "MISSING_FILL",
    "P_LOW_FILL", "P_MID_FILL", "P_HIGH_FILL",
    "PRIMARY_FILL", "PRIMARY_FONT",
    "THIN", "BORDER",
    # 公示番号 → 短縮市区町村名
    "_CITY_CODE_TO_SHORT",
    # ヘルパー関数
    "_format_jpy", "_short_koji_id", "_koji_shape_label", "_short_koji_addr",
    "_format_pct", "_format_hijun_corr", "_hijun_top_bottom",
    "_round_3sig", "_format_price_full",
    "_set", "_section_header", "_adjust_col_widths", "_insert_page_break",
]

# グラフ用の短縮ラベル（表は詳細ラベル、グラフだけ短縮）
SHORT_FEATURE_LABELS = {
    "ln_area": "面積",
    "ln_area_sq": "面積²",
    "walk_min": "駅徒歩",
    "ln_shape": "形状指数",
    "ln_road_w": "道路幅員",
    "ln_far": "容積率",
    "dir_score": "方位",
    "D_shidou": "私道",
    "D_fukuro": "袋地",
    "D_fuseikei": "不整形",
    "ln_district_mean": "地区平均",
    "ln_station_mean": "駅勢圏",
    "const": "定数項",
}

from forbidden_words import assert_clean

# ===== スタイル =====
TITLE_FONT = Font(name="游ゴシック", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="2F5496")
SECTION_FONT = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="4472C4")
LABEL_FONT = Font(name="游ゴシック", size=10, bold=True)
VALUE_FONT = Font(name="游ゴシック", size=10)
BIG_VALUE_FONT = Font(name="游ゴシック", size=18, bold=True, color="C00000")
WARN_FILL = PatternFill("solid", fgColor="FFE699")
MISSING_FILL = PatternFill("solid", fgColor="F4B084")
P_LOW_FILL = PatternFill("solid", fgColor="C6E0B4")  # p<0.05
P_MID_FILL = PatternFill("solid", fgColor="FFE699")  # 0.05<=p<0.1
P_HIGH_FILL = PatternFill("solid", fgColor="F4B084")  # p>=0.1
PRIMARY_FILL = PatternFill("solid", fgColor="E2EFDA")  # 規範性の高い事例の薄緑（モジュール共通）
PRIMARY_FONT = Font(name="游ゴシック", size=10, bold=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _format_jpy(v):
    if v is None:
        return ""
    return f"{int(round(v)):,}円"


_CITY_CODE_TO_SHORT = {
    "13101": "千代田", "13102": "中央", "13103": "港", "13104": "新宿",
    "13105": "文京", "13106": "台東", "13107": "墨田", "13108": "江東",
    "13109": "品川", "13110": "目黒", "13111": "大田", "13112": "世田谷",
    "13113": "渋谷", "13114": "中野", "13115": "杉並", "13116": "豊島",
    "13117": "北", "13118": "荒川", "13119": "板橋", "13120": "練馬",
    "13121": "足立", "13122": "葛飾", "13123": "江戸川",
}


def _short_koji_id(std_id: str) -> str:
    """公示番号を短縮表記。例: '13112-000-050' → '世田谷-50'。"""
    if not std_id:
        return ""
    parts = str(std_id).split("-")
    if len(parts) < 3:
        return str(std_id)
    city_code = parts[0]
    point_num = parts[2]
    city_short = _CITY_CODE_TO_SHORT.get(city_code, city_code)
    try:
        n = int(point_num)
    except (TypeError, ValueError):
        n = point_num
    return f"{city_short}-{n}"


def _koji_shape_label(frontage_ratio, depth_ratio) -> str:
    """公示地点の間口比率(L01_036)・奥行比率(L01_037)から形状ラベルを推定。
    ratio ≤ 1.5 → 整形、≤ 2.5 → やや細長、それ以上 → 細長/不整形。
    """
    if frontage_ratio in (None, "", "_") or depth_ratio in (None, "", "_"):
        return "—"
    try:
        f = float(frontage_ratio)
        d = float(depth_ratio)
    except (TypeError, ValueError):
        return "—"
    if f <= 0 or d <= 0:
        return "—"
    ratio = max(d / f, f / d)
    if ratio <= 1.5:
        return "整形"
    if ratio <= 2.5:
        return "やや細長"
    return "細長"


def _short_koji_addr(address: str, district: str) -> str:
    """公示地点の所在を「地区名+丁目数字」に短縮。
    例: "東京都　世田谷区赤堤５丁目４８４番４" → "赤堤５"
    丁目がない住所は district を返す。
    """
    if not address:
        return district or ""
    if district and district in address:
        after = address[address.index(district):]
        if "丁目" in after:
            return after.split("丁目")[0]
        return district
    # district が住所に見つからない場合は丁目までを返す
    if "丁目" in address:
        # 都道府県・市区町村を除去（最後の "区"/"市"/"町"/"村" 以降）
        for sep in ["区", "市", "町", "村"]:
            if sep in address:
                idx = address.rindex(sep)
                addr_local = address[idx+1:]
                if "丁目" in addr_local:
                    return addr_local.split("丁目")[0]
        return address.split("丁目")[0]
    return district or address


def _format_pct(v):
    if v is None:
        return ""
    return f"{v*100:+.2f}%"


def _format_hijun_corr(multiplier, applies=True, mode="auto"):
    """比準表用の補正値表記（顧客用シート、1セル文字列）。
    mode:
      "top"    = 常に分子側（時点修正：査定時点 / 事例時点）
      "bottom" = 常に分母側（標準化補正・地域格差：100 / 事例評点）
      "auto"   = 補正方向で自動切替
    """
    if not applies:
        return "100/-"
    if abs(multiplier - 1.0) < 0.0005:
        return "100/100"
    if mode == "top":
        return f"{multiplier*100:.1f}/100"
    if mode == "bottom":
        # 上=100, 下=案件評点(=mult*100) — 倍率 = 下/上
        return f"100/{multiplier*100:.1f}"
    if multiplier > 1.0:
        return f"{multiplier*100:.1f}/100"
    return f"100/{multiplier*100:.1f}"


def _hijun_top_bottom(multiplier, applies=True, mode="auto"):
    """比準表用の分子/分母を別々に返す（業者用シートの2行式表示）。
    mode は _format_hijun_corr と同じ。
    Returns: (top, bottom) — 数値または "―"
    """
    if not applies:
        return (100, "―")
    if abs(multiplier - 1.0) < 0.0005:
        return (100, 100)
    if mode == "top":
        return (round(100 * multiplier, 1), 100)
    if mode == "bottom":
        # 上=100, 下=案件評点(=mult*100) — 倍率 = 下/上
        return (100, round(100 * multiplier, 1))
    if multiplier > 1.0:
        return (round(100 * multiplier, 1), 100)
    return (100, round(100 * multiplier, 1))


def _round_3sig(n):
    """上位3桁に四捨五入。例: 424,674,476 → 425,000,000"""
    if n is None or n == 0:
        return n
    import math
    sign = -1 if n < 0 else 1
    n = abs(n)
    digits = int(math.log10(n)) + 1
    if digits <= 3:
        return sign * int(round(n))
    factor = 10 ** (digits - 3)
    return sign * int(round(n / factor) * factor)


def _format_price_full(total_price, area):
    """査定価格表記: 「総額（〇円/㎡、〇円/坪）」を上位3桁四捨五入で。
    坪単価 = ㎡単価 ÷ 0.3025
    """
    if total_price is None or area is None or area <= 0:
        return ""
    total_r = _round_3sig(total_price)
    unit_per_sqm = total_price / area
    unit_per_sqm_r = _round_3sig(unit_per_sqm)
    unit_per_tsubo_r = _round_3sig(unit_per_sqm / 0.3025)
    return f"{total_r:,}円（{unit_per_sqm_r:,}円/㎡、{unit_per_tsubo_r:,}円/坪）"


def _set(ws, row, col, value, font=None, fill=None, align=None, border=False, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = BORDER
    if number_format: c.number_format = number_format
    return c


def _section_header(ws, row, text, end_col=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    _set(ws, row, 1, text, font=SECTION_FONT, fill=SECTION_FILL,
         align=Alignment(horizontal="left", vertical="center"))


def _adjust_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _insert_page_break(ws, before_row):
    """before_row 行の直前で水平改ページを挿入。"""
    from openpyxl.worksheet.pagebreak import Break
    if before_row > 1:
        ws.row_breaks.append(Break(id=before_row - 1))

