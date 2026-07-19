import math
import sys
import json
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from correction import (apply_correction, compute_target_district_mean,
                        compute_target_station_mean, hijun_correction_for_case)
from hedonic import _build_features, annotate_district_mean
from load_mlit import load_mlit_csv
from main import run_pipeline
from time_adjust import annual_rate_for_city


def _hedonic(beta_by_feature):
    return {
        "ok": True,
        "coefficients": {
            name: {"beta": beta, "se": 0.0, "p": 0.0, "label": name}
            for name, beta in beta_by_feature.items()
        },
        "feature_defaults": {"kanguchi": 6.0, "walk_min": 10.0, "road_width": 6.0, "floor_area_ratio": 200.0},
    }


def _case(**kwargs):
    base = {
        "unit_price": 1_000_000.0,
        "adjusted_unit_price": 1_000_000.0,
        "area": math.e,
        "walk_min": 10.0,
        "kanguchi": 6.0,
        "road_width": 6.0,
        "floor_area_ratio": 200.0,
        "road_dir": "北",
        "shape": "整形",
        "road_type": "公道",
    }
    base.update(kwargs)
    return pd.DataFrame([base])


def test_correction_direction_fixed_expected_value():
    target = {"面積(㎡)": math.e**2, "間口": 6.0, "最寄駅:距離(分)": 10}
    actual = apply_correction(_case(), _hedonic({"ln_area": 0.1}), target).iloc[0]["corrected_unit_price"]
    expected = 1_000_000 * math.exp(0.1 * (2 - 1))
    assert actual == pytest.approx(expected)
    assert actual != pytest.approx(1_000_000 / math.exp(0.1 * (2 - 1)))


def test_apply_correction_and_hijun_use_same_price_series():
    target = {"面積(㎡)": math.e**2, "間口": 6.0, "最寄駅:距離(分)": 10}
    corrected = apply_correction(_case(), _hedonic({"ln_area": 0.1}), target)
    assert corrected["canonical_case_price"].equals(corrected["corrected_unit_price"])
    h = hijun_correction_for_case(corrected.iloc[0], _hedonic({"ln_area": 0.1}), target)
    assert h["正本補正後単価"] == pytest.approx(corrected.iloc[0]["corrected_unit_price"])
    assert h["試算値"] == pytest.approx(round(corrected.iloc[0]["corrected_unit_price"], -4), abs=10_000)


def test_canonical_case_price_is_backward_compatible_alias_for_all_rows():
    target = {"面積(㎡)": math.e**2, "間口": 6.0, "最寄駅:距離(分)": 10}
    cases = pd.concat([
        _case(unit_price=1_000_000.0, adjusted_unit_price=1_000_000.0),
        _case(unit_price=1_200_000.0, adjusted_unit_price=1_250_000.0),
    ], ignore_index=True)
    corrected = apply_correction(cases, _hedonic({"ln_area": 0.1}), target)
    assert (corrected["canonical_case_price"] == corrected["corrected_unit_price"]).all()


def test_dir_score_and_fuseikei_are_not_double_counted():
    target = {"面積(㎡)": math.e, "前面道路:方位": "南", "土地の形状": "不整形"}
    hed = _hedonic({"dir_score": 0.1, "D_fuseikei": -0.2})
    actual = apply_correction(_case(road_dir="北", shape="整形"), hed, target).iloc[0]["corrected_unit_price"]
    contrib = 0.1 * (4 - 0) + (-0.2) * (1 - 0)
    assert actual == pytest.approx(1_000_000 * math.exp(contrib))
    assert actual != pytest.approx(1_000_000 * math.exp(2 * contrib))


def test_time_adjustment_evidence_uses_same_two_points_as_rate():
    koji = pd.DataFrame([
        {"標準地番号": "A", "city": "X市", "district": "D", "price_date": date(2024, 1, 1), "price_per_sqm": 100},
        {"標準地番号": "A", "city": "X市", "district": "D", "price_date": date(2025, 1, 1), "price_per_sqm": 110},
        {"標準地番号": "A", "city": "X市", "district": "D", "price_date": date(2026, 1, 1), "price_per_sqm": 220},
    ])
    info = annual_rate_for_city(koji, pd.DataFrame(), "X市", asof=date(2025, 6, 30))
    pt = info["selected_points"][0]
    assert pt["p_prev"] == 100
    assert pt["p_curr"] == 110
    assert pt["date_prev"] == date(2024, 1, 1)
    assert pt["date_curr"] == date(2025, 1, 1)


def test_time_adjustment_does_not_fallback_to_future_data():
    koji = pd.DataFrame([
        {"標準地番号": "A", "city": "X市", "district": "D", "price_date": date(2025, 1, 1), "price_per_sqm": 100},
        {"標準地番号": "A", "city": "X市", "district": "D", "price_date": date(2026, 1, 1), "price_per_sqm": 200},
    ])
    info = annual_rate_for_city(koji, pd.DataFrame(), "X市", asof=date(2025, 6, 30))
    assert info["rate"] is None
    assert info["selected_points"] == []


def test_mlit_loader_excludes_land_with_building(tmp_path):
    path = tmp_path / "mlit.csv"
    pd.DataFrame([
        {"種類": "宅地(土地)", "市区町村名": "X市", "地区名": "D", "取引価格(総額)": 1000, "面積(㎡)": 10, "取引価格(㎡単価)": None, "取引時点": "2025年第1四半期", "前面道路:種類": "市道"},
        {"種類": "宅地(土地と建物)", "市区町村名": "X市", "地区名": "D", "取引価格(総額)": 9999, "面積(㎡)": 10, "取引価格(㎡単価)": None, "取引時点": "2025年第1四半期", "前面道路:種類": "市道"},
    ]).to_csv(path, index=False, encoding="utf-8-sig")
    loaded = load_mlit_csv(path)
    assert len(loaded) == 1
    assert loaded.iloc[0]["unit_price"] == 100


def test_build_features_without_walk_min_column_uses_default():
    X = _build_features(pd.DataFrame([{"area": 100.0}]))
    assert X.iloc[0]["walk_min"] == 10.0


def test_target_encoding_uses_adjusted_price_series_and_leave_one_out():
    df = pd.DataFrame([
        {"district": "D", "station": "S", "unit_price": 100.0, "adjusted_unit_price": 110.0},
        {"district": "D", "station": "S", "unit_price": 200.0, "adjusted_unit_price": 220.0},
        {"district": "D", "station": "S", "unit_price": 300.0, "adjusted_unit_price": 330.0},
    ])
    target = {"地区名": "D", "最寄駅:名称": "S"}
    assert compute_target_district_mean(df, target) == pytest.approx(220.0)
    assert compute_target_station_mean(df, target) == pytest.approx(220.0)
    annotated = annotate_district_mean(df)
    # First row excludes itself: (220 + 330) / 2.
    assert math.exp(annotated.iloc[0]["ln_district_mean"]) == pytest.approx(275.0)


def test_manual_corner_factor_applies_once_when_hedonic_is_skipped():
    hed = {"ok": False, "coefficients": {}}
    target = {"角地補正率(%)": 10}
    corrected = apply_correction(_case(), hed, target)
    assert corrected.iloc[0]["canonical_case_price"] == pytest.approx(1_100_000)
    assert corrected.iloc[0]["corrected_unit_price"] == pytest.approx(1_100_000)


def _cell_after_label(ws, label, col_offset=1):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                return ws.cell(row=cell.row, column=cell.column + col_offset)
    raise AssertionError(f"label not found: {label}")


def _find_cell(ws, label):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                return cell
    raise AssertionError(f"label not found: {label}")


def _cell_before_label(ws, label, row_offset=-1, col_offset=0):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == label:
                return ws.cell(row=cell.row + row_offset, column=cell.column + col_offset)
    raise AssertionError(f"label not found: {label}")


def _parse_unit_from_summary(text):
    return int(text.split("（", 1)[1].split("円/㎡", 1)[0].replace(",", ""))


def test_xlsx_formulas_do_not_reapply_individual_adjustments(tmp_path):
    prop = json.loads((ROOT / "samples" / "sample_property.json").read_text(encoding="utf-8"))
    prop["前面道路:方位"] = "南"
    prop["土地の形状"] = "不整形"
    prop["角地補正率(%)"] = 5
    prop_path = tmp_path / "property_nonzero.json"
    prop_path.write_text(json.dumps(prop, ensure_ascii=False), encoding="utf-8")

    out_path = run_pipeline(
        str(prop_path),
        str(ROOT / "samples" / "sample_mlit.csv"),
        str(ROOT / "samples" / "sample_koji.csv"),
        str(ROOT / "samples" / "sample_kijun.csv"),
        out_dir=str(tmp_path),
        asof=date(2026, 5, 1),
    )
    wb = load_workbook(out_path, data_only=False)
    gyosha = wb["業者用"]
    kokyaku = wb["顧客用"]

    all_formulas = "\n".join(
        str(c.value) for ws in (gyosha, kokyaku) for row in ws.iter_rows()
        for c in row if isinstance(c.value, str) and c.value.startswith("=")
    )
    assert "*(100+B" not in all_formulas
    assert "*(100+C" not in all_formulas
    assert "*B" not in all_formulas
    assert "*C" not in all_formulas
    assert "/C" not in all_formulas

    gyosha_summary_unit = _parse_unit_from_summary(gyosha["A16"].value)
    kokyaku_summary_unit = _parse_unit_from_summary(kokyaku["B3"].value)
    assert gyosha_summary_unit == kokyaku_summary_unit

    gyosha_soan = _cell_after_label(gyosha, "総和")
    assert gyosha_soan.value == 100
    gyosha_anken = _cell_before_label(gyosha, "採用査定単価（top3中央値）")
    assert isinstance(gyosha_anken.value, str) and gyosha_anken.value.startswith("=D")
    gyosha_ref = gyosha[gyosha_anken.value[1:]]
    assert gyosha_ref.value == gyosha_summary_unit

    primary_ref = _cell_after_label(kokyaku, "規範事例の補正後単価（参考）")
    adopted_ref = _cell_after_label(kokyaku, "採用査定単価（top3正本補正後単価の中央値）")
    assert adopted_ref.value == kokyaku_summary_unit

    kokyaku_soan = _cell_after_label(kokyaku, "総和")
    assert kokyaku_soan.value == 100
    kokyaku_anken = _cell_after_label(kokyaku, "案件査定価格（円/㎡）")
    assert isinstance(kokyaku_anken.value, str) and kokyaku_anken.value.startswith("=C")
    kokyaku_ref = kokyaku[kokyaku_anken.value[1:]]
    assert kokyaku_ref.coordinate == adopted_ref.coordinate
    assert kokyaku_ref.value == kokyaku_summary_unit

    # 顧客用の標準化補正・地域格差は、上段=multiplier*100、下段=100。
    for label in ("形状補正", "地域格差"):
        top = _cell_after_label(kokyaku, label)
        bottom = kokyaku.cell(row=top.row + 1, column=top.column)
        assert bottom.value == 100

    # 業者用比準表も標準化補正・地域格差は乗算方向の上段表示。
    for header in ("標準化補正", "地域格差"):
        header_cell = next(c for row in gyosha.iter_rows() for c in row if c.value == header)
        top = gyosha.cell(row=header_cell.row + 1, column=header_cell.column)
        bottom = gyosha.cell(row=header_cell.row + 2, column=header_cell.column)
        assert bottom.value == 100
        assert top.value != "―"


def test_kokyaku_separates_primary_comparable_price_from_adopted_median(tmp_path):
    out_path = run_pipeline(
        str(ROOT / "samples" / "sample_property.json"),
        str(ROOT / "samples" / "sample_mlit.csv"),
        str(ROOT / "samples" / "sample_koji.csv"),
        str(ROOT / "samples" / "sample_kijun.csv"),
        out_dir=str(tmp_path),
        asof=date(2026, 5, 1),
    )
    wb = load_workbook(out_path, data_only=False)
    kokyaku = wb["顧客用"]
    gyosha = wb["業者用"]

    primary_ref = _cell_after_label(kokyaku, "規範事例の補正後単価（参考）")
    adopted_ref = _cell_after_label(kokyaku, "採用査定単価（top3正本補正後単価の中央値）")
    kokyaku_anken = _cell_after_label(kokyaku, "案件査定価格（円/㎡）")
    gyosha_adopted = _cell_before_label(gyosha, "採用査定単価（top3中央値）")

    assert primary_ref.value == 3750000
    assert adopted_ref.value == 3620000
    assert primary_ref.value != adopted_ref.value
    assert kokyaku_anken.value == f"=C{adopted_ref.row}"
    assert gyosha_adopted.value == f"=D{gyosha_adopted.row}"
    assert gyosha[gyosha_adopted.value[1:]].value == adopted_ref.value
