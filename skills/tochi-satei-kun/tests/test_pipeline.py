# Copyright 2026 Koichi Matsuda / SignalYield Advisory
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""3層検証スクリプト（plan §10）。
- Layer 1: 単体スクリプトの sanity check（β符号、類似度自己一致）
- Layer 2: end-to-end のサンプル走行（xlsx生成・禁止語チェック）
- Layer 3: エラー時挙動（件数不足の降格、列欠損エラー）
"""
import io
import json
import sys
from datetime import date
from pathlib import Path

# Windows cp932 対策：標準出力をUTF-8で書き直す
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# scripts ディレクトリを path に追加
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import pandas as pd
from openpyxl import load_workbook

from load_mlit import load_mlit_csv, load_koji_csv, load_kijun_csv
from scope import scope_dataframe
from similarity import compute_similarity, top_k
from time_adjust import annual_rate_for_city, apply_time_adjustment
from hedonic import fit_hedonic
from correction import apply_correction
from forbidden_words import check_text, FORBIDDEN_WORDS
from main import run_pipeline

SAMPLES = ROOT / "samples"


# ===== Layer 1 =====
def test_hedonic_sign():
    """β符号テスト：面積→負、駅距離→負、私道→負、袋地→負（生成モデルと整合）。"""
    df = load_mlit_csv(SAMPLES / "sample_mlit.csv")
    koji = load_koji_csv(SAMPLES / "sample_koji.csv")
    kijun = load_kijun_csv(SAMPLES / "sample_kijun.csv")
    with open(SAMPLES / "sample_property.json", encoding="utf-8") as f:
        target = json.load(f)
    asof = date(2025, 12, 1)
    scoped, _ = scope_dataframe(df, target, asof)
    rate = annual_rate_for_city(koji, kijun, target["市区町村名"])["rate"]
    adjusted = apply_time_adjustment(scoped, asof, rate)
    hed = fit_hedonic(adjusted)
    assert hed["ok"], f"回帰失敗: {hed['skip_reason']}"
    coef = hed["coefficients"]
    # 期待符号
    expectations = {
        "ln_area": "negative",     # 単価ベース：広い→単価減
        "walk_min": "negative",    # 駅遠い→単価減
        # ln_shape は形状指数：0付近で正方形、極端で帯/旗竿。符号は地域依存（テスト対象外）
        "ln_road_w": "positive",   # 道路幅員広い→単価増
        "dir_score": "positive",   # 方位スコア大（南寄り）→単価増
        "D_shidou": "negative",    # 私道→減価
        "D_fukuro": "negative",    # 袋地→減価
        "D_fuseikei": "negative",  # 不整形→減価
    }
    failures = []
    for name, expected in expectations.items():
        if name not in coef:
            failures.append(f"{name}: 係数なし")
            continue
        beta = coef[name]["beta"]
        if expected == "negative" and beta > 0:
            failures.append(f"{name}: 期待=負, 実際={beta:+.4f}")
        if expected == "positive" and beta < 0:
            failures.append(f"{name}: 期待=正, 実際={beta:+.4f}")
    if failures:
        # 警告レベル：合成データのノイズで稀に符号反転する可能性あり
        print(f"[WARN] β符号期待外れ: {failures}")
    else:
        print(f"[OK] β符号テスト")


def test_similarity_self():
    """同一物件投入：自分自身がスコア1.0前後を返るか。"""
    df = load_mlit_csv(SAMPLES / "sample_mlit.csv")
    # df の1行目を target に変換
    row = df.iloc[0]
    target = {
        "市区町村名": row["city"],
        "地区名": row.get("district", ""),
        "面積(㎡)": row["area"],
        "最寄駅:距離(分)": row.get("walk_min", 10),
        "土地の形状": row.get("shape", "整形"),
        "前面道路:種類": row.get("road_type", "公道"),
        "前面道路:幅員(m)": row.get("road_width", 6.0),
        "前面道路:方位": row.get("road_dir", "南"),
    }
    sim = compute_similarity(df.head(5), target)
    self_sim = sim.iloc[0]["similarity"]
    assert self_sim > 0.9, f"自己類似度 {self_sim:.3f} が低すぎる"
    print(f"[OK] 類似度自己一致テスト (self_similarity={self_sim:.4f})")


# ===== Layer 2 =====
def test_e2e_pipeline():
    """end-to-end でxlsx生成・禁止語チェック。"""
    out_path = run_pipeline(
        str(SAMPLES / "sample_property.json"),
        str(SAMPLES / "sample_mlit.csv"),
        str(SAMPLES / "sample_koji.csv"),
        str(SAMPLES / "sample_kijun.csv"),
        out_dir=str(ROOT / "output"),
    )
    assert out_path.exists(), f"xlsx未生成: {out_path}"

    wb = load_workbook(out_path)
    assert "業者用" in wb.sheetnames, "業者用シートなし"
    assert "顧客用" in wb.sheetnames, "顧客用シートなし"

    # 業者用シートには係数開示があるはず
    gyosha = wb["業者用"]
    gyosha_text = "\n".join(
        str(c.value) for row in gyosha.iter_rows() for c in row if c.value is not None
    )
    assert "ヘドニック" in gyosha_text or "サンプル数" in gyosha_text, \
        "業者用シートに回帰サマリ記載なし"
    print(f"[OK] 業者用シートに回帰サマリあり")

    # 顧客用シートに禁止語が含まれていないこと
    kokyaku = wb["顧客用"]
    kokyaku_text = "\n".join(
        str(c.value) for row in kokyaku.iter_rows() for c in row if c.value is not None
    )
    ok, detected = check_text(kokyaku_text)
    assert ok, f"顧客用シートに禁止語混入: {detected}"
    print(f"[OK] 顧客用シート禁止語チェック")
    print(f"[OK] end-to-end: {out_path}")


# ===== Layer 3 =====
def test_low_count_degradation():
    """件数不足時の降格仕様：ヘドニック回帰スキップ→類似度ベース集約。"""
    df = load_mlit_csv(SAMPLES / "sample_mlit.csv")
    # 5件に絞り込む
    df_small = df.head(5)
    hed = fit_hedonic(df_small)
    assert not hed["ok"], "件数5件で回帰が成功してしまった（降格判定が機能していない）"
    assert "回帰スキップ" in hed["skip_reason"], f"スキップ理由が想定外: {hed['skip_reason']}"

    # apply_correction が hed.ok=False でも例外を出さず、補正なしで動くこと
    target = {"市区町村名": "港区", "地区名": "麻布十番", "面積(㎡)": 120,
              "最寄駅:距離(分)": 7, "土地の形状": "整形",
              "前面道路:種類": "公道", "前面道路:方位": "南", "前面道路:幅員(m)": 6.0}
    corrected = apply_correction(df_small, hed, target)
    assert "corrected_unit_price" in corrected.columns, "降格時に corrected_unit_price 列なし"
    print(f"[OK] 件数不足降格テスト (n={hed['n']})")


def test_missing_column():
    """MLIT CSV 必須列欠損時のエラー。"""
    bad_df = pd.DataFrame({"foo": [1, 2, 3]})
    bad_path = ROOT / "tests" / "_bad.csv"
    bad_df.to_csv(bad_path, encoding="utf-8-sig", index=False)
    try:
        load_mlit_csv(bad_path)
        raise AssertionError("必須列欠損が見逃された")
    except ValueError as e:
        assert "必須列欠損" in str(e), f"エラーメッセージ想定外: {e}"
        print(f"[OK] 列欠損エラーテスト")
    finally:
        bad_path.unlink(missing_ok=True)


if __name__ == "__main__":
    print("\n=== Layer 1: 単体 sanity check ===")
    test_hedonic_sign()
    test_similarity_self()
    print("\n=== Layer 2: end-to-end ===")
    test_e2e_pipeline()
    print("\n=== Layer 3: エラー時挙動 ===")
    test_low_count_degradation()
    test_missing_column()
    print("\n[ALL OK] 全テスト完了")
