# Copyright 2026 Koichi Matsuda / SignalYield Advisory
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""xlsx 生成のオーケストレータ（v1.2.7 で分割：xlsx_common / xlsx_gyosha_sheet / xlsx_kokyaku_sheet）。

write_xlsx() がエントリポイント。各シート描画は専用モジュールに委譲。

確定方針（プラン §7）：
- 業者用：n, R², 全β、補正内訳、参考値、警告
- 顧客用：ですます調、禁止語ブロック
"""
from pathlib import Path

from openpyxl import Workbook

from xlsx_common import *
from xlsx_gyosha_sheet import _write_gyosha_sheet
from xlsx_kokyaku_sheet import _write_kokyaku_sheet
from version import ENGINE_VERSION


def _apply_page_setup(wb: Workbook, target: dict):
    """各シートに印刷設定を適用。
    業者用: A3 横 + 縮小印刷（fitToWidth=1）
    グラフ: A4 横 + 縮小印刷（fitToPage）
    顧客用: A4 縦 + ヘッダ「机上査定書」+ フッタ「N / 総页」
    """
    from openpyxl.worksheet.page import PageMargins
    # PAPERSIZE: A3=8, A4=9
    A3_SIZE, A4_SIZE = 8, 9
    target_label = target.get("物件略号", "")
    location = f"{target.get('市区町村名','')} {target.get('地区名','')}{target.get('丁目','')}"

    # 業者用：A3 横、幅に合わせて縮小、印刷タイトルとして 1行目を固定
    if "業者用" in wb.sheetnames:
        ws = wb["業者用"]
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = A3_SIZE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 縦は複数ページ可
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6,
                                       header=0.3, footer=0.3)
        ws.print_title_rows = "1:1"
        ws.oddHeader.left.text = f"業者用 — {target_label}"
        ws.oddHeader.right.text = "&D"
        ws.oddFooter.center.text = "&P / &N"
        ws.sheet_view.view = "pageBreakPreview"

    # グラフ：A4 横、ページに合わせて縮小
    if "グラフ" in wb.sheetnames:
        gs = wb["グラフ"]
        gs.page_setup.orientation = gs.ORIENTATION_LANDSCAPE
        gs.page_setup.paperSize = A4_SIZE
        gs.page_setup.fitToWidth = 1
        gs.page_setup.fitToHeight = 1
        gs.sheet_properties.pageSetUpPr.fitToPage = True
        gs.print_options.horizontalCentered = True
        gs.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6)
        gs.oddHeader.left.text = f"附属資料 — {target_label}"
        gs.oddFooter.center.text = "&P / &N"
        gs.sheet_view.view = "pageBreakPreview"

    # 顧客用：A4 縦、ヘッダ「机上査定書」、フッタ「現/総」
    if "顧客用" in wb.sheetnames:
        ks = wb["顧客用"]
        ks.page_setup.orientation = ks.ORIENTATION_PORTRAIT
        ks.page_setup.paperSize = A4_SIZE
        ks.page_setup.fitToWidth = 1
        ks.page_setup.fitToHeight = 0
        ks.sheet_properties.pageSetUpPr.fitToPage = True
        ks.print_options.horizontalCentered = True
        ks.page_margins = PageMargins(left=0.6, right=0.6, top=0.8, bottom=0.8,
                                       header=0.3, footer=0.4)
        ks.oddHeader.center.text = "&\"游ゴシック,Bold\"&14机上査定書"
        ks.oddHeader.right.text = location
        ks.oddFooter.center.text = "&P / &N"
        ks.sheet_view.view = "pageBreakPreview"


def write_xlsx(ctx: dict, output_path: Path) -> Path:
    wb = Workbook()
    # Workbook プロパティに認証情報を埋め込む（Excel「ファイル → 情報 → プロパティ」で確認可能、
    # ハルシネーション出力との判別用）
    wb.properties.creator = f"tochi-satei-kun v{ENGINE_VERSION}"
    wb.properties.description = (
        "土地価格査定クン (tochi-satei-kun) — Apache License 2.0 OSS AVM. "
        "https://github.com/signal-yield/tochi-satei-kun"
    )
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _write_gyosha_sheet(wb, ctx)
    _write_kokyaku_sheet(wb, ctx)
    _apply_page_setup(wb, ctx.get("target", {}))
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
